from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List, Dict
from database.db_config import get_db
from database.models import Product, Review, AnalysisResult
from datetime import datetime
from services.scraper import scraper
from services.sentiment_analyzer import sentiment_analyzer
import json
import csv
import io
from openpyxl import load_workbook

router = APIRouter()

class ProductRequest(BaseModel):
    url: str
    platform: Optional[str] = "mercadolibre"
    name: Optional[str] = None

class ProductResponse(BaseModel):
    id: int
    name: str
    platform: str
    url: str
    image_url: Optional[str] = None
    price: Optional[float] = None
    created_at: datetime
    
    class Config:
        from_attributes = True

@router.post("/search", response_model=List[dict])
async def search_products(
    product_name: str,
    platforms: Optional[List[str]] = None
):
    """
    Búsqueda de productos por nombre.
    Actualmente no implementado para múltiples plataformas; se recomienda usar URL directa de Mercado Libre.
    """
    if not product_name or not product_name.strip():
        raise HTTPException(status_code=400, detail="Product name is required")
    
    # Deshabilitado: la búsqueda multi-plataforma fue retirada para centrarse en Mercado Libre.
    raise HTTPException(status_code=501, detail="Search by name is not implemented. Use Mercado Libre product URL.")

@router.post("/", response_model=ProductResponse)
async def create_product(product: ProductRequest, db: Session = Depends(get_db)):
    """
    Create a new product entry for analysis
    """
    # Check if product already exists
    existing = db.query(Product).filter(Product.url == product.url).first()
    if existing:
        return existing
    
    # Create new product
    db_product = Product(
        name=product.name or "Unknown Product",
        platform=product.platform,
        url=product.url
    )
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product

@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(product_id: int, db: Session = Depends(get_db)):
    """
    Get product details by ID
    """
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product

@router.get("/", response_model=List[ProductResponse])
async def list_products(skip: int = 0, limit: int = 10, db: Session = Depends(get_db)):
    """
    List all analyzed products
    """
    products = db.query(Product).offset(skip).limit(limit).all()
    return products

@router.delete("/{product_id}")
async def delete_product(product_id: int, db: Session = Depends(get_db)):
    """
    Delete a product and all its associated data
    """
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    db.delete(product)
    db.commit()
    return {"message": "Product deleted successfully"}


class UploadAnalysisResponse(BaseModel):
    product_id: int
    product_name: str
    stars: float
    sentiment_label: str
    avg_sentiment: float
    total_reviews: int
    positive_count: int
    neutral_count: int
    negative_count: int
    keywords: List[str]
    opinion_summary: str

def _summarize(analysis: Dict) -> Dict:
    avg = float(analysis.get("avg_sentiment", 0.5))
    label = str(analysis.get("sentiment_label", "neutral"))
    total = int(analysis.get("total_reviews", 0))
    keywords = list(analysis.get("keywords", []))
    # Ajuste de estrellas: reflejar fielmente el promedio (0–5) sin imponer mínimo 1
    stars = round(max(0.0, min(5.0, avg * 5.0)), 1)
    summary_base = {
        "positive": "Opiniones mayormente positivas.",
        "negative": "Opiniones mayormente negativas.",
        "neutral": "Opiniones mixtas o neutrales.",
    }.get(label, "Opiniones mixtas o neutrales.")
    # Distribución porcentual si está disponible
    dist = analysis.get("sentiment_distribution", {}) or {}
    pos = dist.get("positive")
    neu = dist.get("neutral")
    neg = dist.get("negative")
    parts: List[str] = [summary_base]
    parts.append(f"Promedio {stars}★ con {total} reseñas.")
    if isinstance(pos, (int, float)) and isinstance(neu, (int, float)) and isinstance(neg, (int, float)):
        parts.append(f"Distribución: {pos}% positivas, {neu}% neutrales, {neg}% negativas.")
    if keywords:
        parts.append(f"Temas destacados: {', '.join(keywords[:6])}.")
    opinion_summary = " ".join(parts)
    return {
        "stars": stars,
        "sentiment_label": label,
        "avg_sentiment": round(avg, 3),
        "total_reviews": total,
        "keywords": keywords,
        "opinion_summary": opinion_summary,
    }

def _parse_file(file: UploadFile) -> List[Dict]:
    name = (file.filename or "").lower()
    data: List[Dict] = []
    content = file.file.read()
    if name.endswith(".json"):
        try:
            items = json.loads(content.decode("utf-8"))
            if isinstance(items, list):
                data = items
        except Exception:
            pass
    elif name.endswith(".csv"):
        try:
            text = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                data.append(row)
        except Exception:
            pass
    elif name.endswith(".xlsx"):
        try:
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                data.append(item)
        except Exception:
            pass
    else:
        # Intento básico: probar JSON
        try:
            items = json.loads(content.decode("utf-8"))
            if isinstance(items, list):
                data = items
        except Exception:
            pass
    return data

@router.post("/upload", response_model=UploadAnalysisResponse)
async def upload_reviews_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Recibe un archivo (.json, .csv, .xlsx) con reseñas y realiza análisis de sentimientos.
    Campos esperados (flexibles): text, rating, product_name.
    Persiste reseñas bajo un producto "Uploaded Dataset" para pruebas locales.
    """
    rows = _parse_file(file)
    if not rows:
        raise HTTPException(status_code=400, detail="No data parsed from file")

    product_name = None
    reviews: List[Dict] = []
    for r in rows:
        text = (r.get("text") or r.get("review_text") or "").strip()
        rating = r.get("rating")
        try:
            rating = float(rating) if rating is not None else None
        except Exception:
            rating = None
        product_name = product_name or (r.get("product_name") or r.get("item_name") or "Uploaded Dataset")
        reviews.append({"text": text, "rating": rating})

    # Crear/obtener producto (idempotente por URL)
    file_url = f"file://{file.filename}"
    existing = db.query(Product).filter(Product.url == file_url).first()
    if existing:
        db_product = existing
        # Limpiar reseñas y análisis previos para evitar duplicados al re-subir
        db.query(Review).filter(Review.product_id == db_product.id).delete(synchronize_session=False)
        db.query(AnalysisResult).filter(AnalysisResult.product_id == db_product.id).delete(synchronize_session=False)
        db.commit()
    else:
        db_product = Product(name=product_name or "Uploaded Dataset", platform="dataset", url=file_url)
        db.add(db_product)
        db.commit()
        db.refresh(db_product)

    # Guardar reseñas
    for r in reviews:
        db.add(Review(
            product_id=db_product.id,
            user_name="Anonymous",
            rating=(r.get("rating") or 3.0),
            text=r.get("text") or "",
            review_date=datetime.utcnow(),
            platform="dataset",
        ))
    db.commit()

    # Análisis
    analysis = sentiment_analyzer.analyze_reviews(reviews)
    summary = _summarize(analysis)

    # Persistir resultado de análisis resumido
    db.add(AnalysisResult(
        product_id=db_product.id,
        avg_sentiment=summary["avg_sentiment"],
        sentiment_label=summary["sentiment_label"],
        total_reviews=summary["total_reviews"],
        positive_count=analysis.get("positive_count", 0),
        negative_count=analysis.get("negative_count", 0),
        neutral_count=analysis.get("neutral_count", 0),
        keywords=summary["keywords"],
        price_data=None,
        analyzed_at=datetime.utcnow(),
    ))
    db.commit()

    return UploadAnalysisResponse(
        product_id=db_product.id,
        product_name=db_product.name,
        stars=summary["stars"],
        sentiment_label=summary["sentiment_label"],
        avg_sentiment=summary["avg_sentiment"],
        total_reviews=summary["total_reviews"],
        positive_count=analysis.get("positive_count", 0),
        neutral_count=analysis.get("neutral_count", 0),
        negative_count=analysis.get("negative_count", 0),
        keywords=summary["keywords"],
        opinion_summary=summary["opinion_summary"],
    )
